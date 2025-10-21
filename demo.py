from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import data_py
data_py.load_users()


def export_week(data1):
    import data_py
    t=data_py.team.read_mainfile()
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Process each team in a separate sheet
    for team_id, team_data in data1.items():
        for i in t["idteam"]:
            if i["id_team"] == team_id:
                name=str(i["name"])[0].upper()+str(i["name"])[1:]
        # Create new sheet for team
        ws = wb.create_sheet(name)
        
        # Styles
        header_font = Font(bold=True)
        header_fill = PatternFill(fill_type='solid', start_color='CCE5FF')
        center_align = Alignment(horizontal='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))

        # Headers
        headers = ['Mã học sinh', 'Họ và tên', 'Điểm cộng', 'Điểm trừ', 'Tổng điểm', 'Xếp loại', 'Hạng']
        
        # Sort students by total points for ranking
        sorted_students = sorted(
            [(student_id, student) for student_id, student in team_data.items()],
            key=lambda x: x[1]['total'],
            reverse=True
        )
        
        # Assign ranks
        base_rank = 1
        prev_total = None
        ranks = {}
        for student_id, student in sorted_students:
            if prev_total is None or student['total'] != prev_total:
                ranks[student_id] = base_rank
            else:
                ranks[student_id] = ranks[prev_student_id]
            base_rank += 1
            prev_total = student['total']
            prev_student_id = student_id

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
            
        # Write student data
        current_row = 2
        for idx, (student_id, student) in enumerate(team_data.items(), 1):
            # Mã học sinh
            ws.cell(row=current_row, column=1, value=student_id)
            # Student data
            ws.cell(row=current_row, column=2, value=student['name'])
            ws.cell(row=current_row, column=3, value=student['give'])
            ws.cell(row=current_row, column=4, value=student['error'])
            ws.cell(row=current_row, column=5, value=student['total'])
            ws.cell(row=current_row, column=6, value=student['ratings'])
            ws.cell(row=current_row, column=7, value=ranks[student_id])
            
            # Style cells
            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.alignment = center_align
                cell.border = border
            
            current_row += 1

        # Auto-adjust columns for this sheet
        for col in range(1, 8):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in ws.rows:
                cell = row[col-1]
                try:
                    if cell.value:
                        max_length = max(len(str(cell.value)), max_length)
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

    wb.save('week.xlsx')
    return 'week.xlsx'

def export_semester(data1):
    import data_py
    t = data_py.team.read_mainfile()

    wb = Workbook()
    wb.remove(wb.active)

    # Xử lý từng team
    for team_id, students in data1.items():
        # Lấy tên team
        for i in t["idteam"]:
            if i["id_team"] == team_id:
                name = str(i["name"])[0].upper() + str(i["name"])[1:]
        # print(students)
        # Tạo sheet mới
        ws = wb.create_sheet(name)

        # Style cơ bản
        header_font = Font(bold=True)
        header_fill = PatternFill(fill_type='solid', start_color='CCE5FF')
        center_align = Alignment(horizontal='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Header
        headers = ['STT', 'Họ và tên', 'Tổng điểm', 'Xếp loại', 'Hạng']

        # Sắp xếp học sinh theo điểm giảm dần
        sorted_students = sorted(students, key=lambda s: s[1], reverse=True)

        # Gán hạng (có xử lý đồng hạng)
        ranks = []
        prev_score = None
        rank = 0
        for idx, s in enumerate(sorted_students, 1):
            if s[1] != prev_score:
                rank = idx
            ranks.append(rank)
            prev_score = s[1]

        # Ghi header
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Ghi dữ liệu học sinh
        for row_idx, (student, rank) in enumerate(zip(sorted_students, ranks), start=2):
            name, score, rating = student
            ws.cell(row=row_idx, column=1, value=data.find_user_name(name)["id"])
            ws.cell(row=row_idx, column=2, value=name)
            ws.cell(row=row_idx, column=3, value=score)
            ws.cell(row=row_idx, column=4, value=rating)
            ws.cell(row=row_idx, column=5, value=rank)

            for col in range(1, 6):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = center_align
                cell.border = border

        # Auto chỉnh độ rộng cột
        for col in range(1, 6):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            ws.column_dimensions[column_letter].width = max_length + 2

    wb.save('semester.xlsx')
    return 'semester.xlsx'

def export_year(data1):
    import data_py
    t = data_py.team.read_mainfile()

    wb = Workbook()
    wb.remove(wb.active)
    # Xử lý từng team
    for team_id, students in data1.items():
        # Lấy tên team
        for i in t["idteam"]:
            if i["id_team"] == team_id:
                name = str(i["name"])[0].upper() + str(i["name"])[1:]

        # Tạo sheet mới
        ws = wb.create_sheet(name)
        students=students["students"]
        # Style cơ bản
        header_font = Font(bold=True)
        header_fill = PatternFill(fill_type='solid', start_color='CCE5FF')
        center_align = Alignment(horizontal='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Header
        headers = ['STT', 'Họ và tên', 'Tổng điểm', 'Xếp loại', 'Hạng']

        # Sắp xếp học sinh theo điểm giảm dần
        sorted_students = sorted(students, key=lambda s: s[1], reverse=True)

        # Gán hạng (có xử lý đồng hạng)
        ranks = []
        prev_score = None
        rank = 0
        for idx, s in enumerate(sorted_students, 1):
            if s[1] != prev_score:
                rank = idx
            ranks.append(rank)
            prev_score = s[1]

        # Ghi header
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Ghi dữ liệu học sinh
        for row_idx, (student, rank) in enumerate(zip(sorted_students, ranks), start=2):
            name, score, rating = student
            ws.cell(row=row_idx, column=1, value=data.find_user_name(name)["id"])
            ws.cell(row=row_idx, column=2, value=name)
            ws.cell(row=row_idx, column=3, value=score)
            ws.cell(row=row_idx, column=4, value=rating)
            ws.cell(row=row_idx, column=5, value=rank)

            for col in range(1, 6):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = center_align
                cell.border = border

        # Auto chỉnh độ rộng cột
        for col in range(1, 6):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            ws.column_dimensions[column_letter].width = max_length + 2

    wb.save('year.xlsx')
    return 'year.xlsx'

# Test data
week = {
    '2': {
        "3": {'name': 'thanh', 'give': 0, 'error': 0, 'total': 100, 'ratings': 'Super good'},
        "7": {'name': 'demo3', 'give': 0, 'error': 0, 'total': 150, 'ratings': 'Super good'}  # Higher total
    },
    '5': {
        6: {'name': 'demo2', 'give': 0, 'error': 0, 'total': 100, 'ratings': 'Super good'},
        8: {'name': 'demo4', 'give': 0, 'error': 0, 'total': 100, 'ratings': 'Super good'}  # Equal totals
    }
}

semester = {
    '2': [["thanh", 1800, "Super good"], ["demo3", 1785, "Super good"]],
    '5': [["demo2", 23572, "Super good"], ["demo4", 1800, "Super good"]]
}

year={
    "2": {"students": [["thanh",1700,"Super good"],["demo3",1715,"Super good"]]},
    "5": {"students": [["demo2",1709,"Super good"],["demo4",1700,"Super good"]]}
}

# export_week(week)
# export_semester(semester)
# export_year(year)

# import config
# t2=data.summary.read_main("week")

# for h in range(config.semester_1+1,t2["num"]+1):
#     print(h)

# import data_py

t1=data_py.summary.read_main("semester")
t3=data_py.team.read_mainfile()
t5=data_py.summary.read(1, "year", 1)
    # print(t5)