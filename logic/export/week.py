from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import data


def export_week(data1, id_class):
    t = data.team.read_mainfile(id_class)
    wb = Workbook()
    wb.remove(wb.active)

    for team_id, team_data in data1.items():
        name = ""
        for i in t["idteam"]:
            if str(i["id_team"]) == team_id:
                name = str(i["name"])[0].upper() + str(i["name"])[1:]
        if not name:
            name = f"Team_{team_id}"

        ws = wb.create_sheet(name)

        header_font = Font(bold=True)
        header_fill = PatternFill(fill_type='solid', start_color='CCE5FF')
        center_align = Alignment(horizontal='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['Mã học sinh', 'Họ và tên', 'Điểm cộng', 'Điểm trừ', 'Tổng điểm', 'Xếp loại', 'Hạng']

        students_only = {k: v for k, v in team_data.items() if isinstance(v, dict)}
        if students_only:
            sorted_students = sorted(
                [(student_id, student) for student_id, student in students_only.items()],
                key=lambda x: x[1]['total'],
                reverse=True
            )
        else:
            sorted_students = []

        base_rank = 1
        prev_total = None
        ranks = {}
        prev_student_id = None
        for student_id, student in sorted_students:
            if prev_total is None or student['total'] != prev_total:
                ranks[student_id] = base_rank
            else:
                ranks[student_id] = ranks[prev_student_id]
            base_rank += 1
            prev_total = student['total']
            prev_student_id = student_id

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        current_row = 2
        for idx, (student_id, student) in enumerate(students_only.items(), 1):
            ws.cell(row=current_row, column=1, value=student_id)
            ws.cell(row=current_row, column=2, value=student['name'])
            ws.cell(row=current_row, column=3, value=student['give'])
            ws.cell(row=current_row, column=4, value=student['error'])
            ws.cell(row=current_row, column=5, value=student['total'])
            ws.cell(row=current_row, column=6, value=student['ratings'])
            ws.cell(row=current_row, column=7, value=ranks.get(student_id, idx))

            for col in range(1, 8):
                cell = ws.cell(row=current_row, column=col)
                cell.alignment = center_align
                cell.border = border

            current_row += 1

        for col in range(1, 8):
            max_length = 0
            column_letter = get_column_letter(col)
            for row in ws.rows:
                cell = row[col - 1]
                try:
                    if cell.value:
                        max_length = max(len(str(cell.value)), max_length)
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
