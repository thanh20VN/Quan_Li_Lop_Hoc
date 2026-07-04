from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import data_py


def export_year(data1, id_class):
    t = data_py.team.read_mainfile(id_class)
    wb = Workbook()
    wb.remove(wb.active)

    for team_id, students_data in data1.items():
        name = ""
        for i in t["idteam"]:
            if str(i["id_team"]) == str(team_id):
                name = str(i["name"])[0].upper() + str(i["name"])[1:]
        if not name:
            name = f"Team_{team_id}"

        ws = wb.create_sheet(name)

        if isinstance(students_data, dict):
            students = students_data.get("students", [])
        else:
            students = students_data

        header_font = Font(bold=True)
        header_fill = PatternFill(fill_type='solid', start_color='CCE5FF')
        center_align = Alignment(horizontal='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = ['STT', 'Họ và tên', 'Tổng điểm', 'Xếp loại', 'Hạng']

        if students:
            sorted_students = sorted(students, key=lambda s: s[1], reverse=True)
        else:
            sorted_students = []

        ranks = []
        prev_score = None
        rank = 0
        for idx, s in enumerate(sorted_students, 1):
            if s[1] != prev_score:
                rank = idx
            ranks.append(rank)
            prev_score = s[1]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        for row_idx, (student, rank) in enumerate(zip(sorted_students, ranks), start=2):
            name_str, score, rating = student
            user = data_py.find_user_name(name_str)
            user_id = user["id"] if isinstance(user, dict) else 0
            ws.cell(row=row_idx, column=1, value=user_id)
            ws.cell(row=row_idx, column=2, value=name_str)
            ws.cell(row=row_idx, column=3, value=score)
            ws.cell(row=row_idx, column=4, value=rating)
            ws.cell(row=row_idx, column=5, value=rank)

            for col in range(1, 6):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = center_align
                cell.border = border

        for col in range(1, 6):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            ws.column_dimensions[column_letter].width = max_length + 2

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
