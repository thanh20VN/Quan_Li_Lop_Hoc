from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import data_py

def __init__(data1):
    t = data_py.team.read_mainfile()

    wb = Workbook()
    wb.remove(wb.active)

    # Xử lý từng team
    for team_id, students in data1.items():
        # Lấy tên team
        for i in t["idteam"]:
            if str(i["id_team"]) == team_id:
                name = str(i["name"])[0].upper() + str(i["name"])[1:]
        # print(students)
        # Tạo sheet mới
        ws = wb.create_sheet(name)

        # Style cơ bản
        header_font = Font(bold=True)
        header_fill = PatternFill(fill_type='solid', start_color='CCE5FF')
        center_align = Alignment(horizontal='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Header
        headers = ['STT', 'Họ và tên', 'Tổng điểm', 'Xếp loại', 'Hạng']

        # Sắp xếp học sinh theo điểm giảm dần
        sorted_students = sorted(students, key=lambda s: s[1], reverse=True)

        # Gán hạng (có xử lý đồng hạng)
        ranks = []
        prev_score = None
        rank = 0
        for idx, s in enumerate(sorted_students, 1):
            if s[1] != prev_score:
                rank = idx
            ranks.append(rank)
            prev_score = s[1]

        # Ghi header
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border

        # Ghi dữ liệu học sinh
        for row_idx, (student, rank) in enumerate(zip(sorted_students, ranks), start=2):
            name, score, rating = student
            ws.cell(row=row_idx, column=1, value=data_py.find_user_name(name)["id"])
            ws.cell(row=row_idx, column=2, value=name)
            ws.cell(row=row_idx, column=3, value=score)
            ws.cell(row=row_idx, column=4, value=rating)
            ws.cell(row=row_idx, column=5, value=rank)

            for col in range(1, 6):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = center_align
                cell.border = border

        # Auto chỉnh độ rộng cột
        for col in range(1, 6):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            ws.column_dimensions[column_letter].width = max_length + 2

    wb.save('semester.xlsx')
    return 'semester.xlsx'
